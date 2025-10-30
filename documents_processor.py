# -*- coding: utf-8 -*-
"""
Universal document processor extracting text and images from various formats.
Now aligned with local 'document_processor_rar_zip.py':
- Optional direct Excel processing via ENABLE_DIRECT_EXCEL (default: off)
- Safe Excel utilities for .xlsx/.xls
- PDF page preview at dpi=200, with unique image filenames
- AI Vision orientation fix before description
"""

import os
import io
import logging
import shutil
import base64
import zipfile
# Optional dependency for .rar support
try:
    import rarfile  # needs `pip install rarfile` and unrar/bsdtar on system
except ImportError:
    rarfile = None
import subprocess
from datetime import datetime
from typing import List, Dict, Optional
from pathlib import Path
import json
import email
from email import policy
from email.parser import BytesParser
# Optional HTML->text conversion
try:
    from bs4 import BeautifulSoup  # type: ignore
except Exception:
    BeautifulSoup = None  # type: ignore

import fitz  # PyMuPDF
import pandas as pd
from PIL import Image, ExifTags
from PIL import ImageOps  # ← for auto-rotating images by EXIF
from striprtf.striprtf import rtf_to_text
import strip_markdown
import openpyxl
from functools import lru_cache
from dotenv import load_dotenv

# Legacy .xls optional support
try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False

from openai import OpenAI   # ⬅️ new-style SDK (≥ 1.0.0)  ✅

# Optional dependency for Apple Numbers (.numbers)
try:
    from numbers_parser import Document as NumbersDocument  # type: ignore
    NUMBERS_PARSER_AVAILABLE = True
except Exception:
    NumbersDocument = None  # type: ignore
    NUMBERS_PARSER_AVAILABLE = False

# Module-level AI instruction for image description
AI_IMAGE_DESCRIPTION_PROMPT = """You are given a single image that may depict a document or a general scene. Analyze the image and provide a detailed report with the following structure:

1. General Description
   - Briefly describe what is visible (layout, key elements, objects)
   - If people are present, describe the scene neutrally (e.g., clothing, actions, count, surroundings) without identifying individuals or inferring sensitive attributes. Do not perform face recognition or unique identification.
   - If the document/form is empty or contains no meaningful information, note this and keep the summary concise.

2. Document Type Classification
   - If applicable, identify the document type.
   - Otherwise, classify as "Other document type" (e.g., photograph, scene).

3. Text Content (if applicable)
   - Extract visible text exactly as it appears (if any).
   - Maintain the original order of paragraphs and lines.
   - Do not omit or paraphrase meaningful information (e.g., filled form fields and values).

4. Fields and Values (for documents)
   - For each labeled field, heading, or identifier, specify:
     • Field name (as labeled in the document)
     • Purpose/meaning (brief explanation)

5. Element Summary
   - Provide a comprehensive list of all mentioned elements, fields, and sections.
   - Include exact wording where present and brief context.

Notes:
- Keep the description neutral and factual. Avoid identifying individuals or inferring sensitive attributes.
- If the content is not a document, still provide a useful scene description following the structure above."""

# ----------------------------------------------------------------------
# Configuration constants
# ----------------------------------------------------------------------
MIN_IMG_PIXELS = 200 * 200
MAX_VISION_CALLS_PER_PAGE = 50
MAX_IMAGE_DIM = 3000
MAX_IMAGE_SIZE = 10 * 1024 * 1024

SUPPORTED_IMAGE_FORMATS = {".jpg", ".jpeg", ".png", ".heic", ".heif", ".gif", ".tiff", ".tif", ".bmp"}

# Compact preview and heuristic thresholds
CSV_PREVIEW_CHARS = 1000
HTML_IMAGE_HEURISTIC_BYTES = 10000

# Allowed extensions inside archives (ZIP/RAR)
ARCHIVE_ALLOWED_EXTS = SUPPORTED_IMAGE_FORMATS | {
    ".pdf", ".txt", ".pages", ".numbers", ".xlsx", ".xls", ".csv", ".json", ".py",
    ".html", ".cms", ".css", ".eml", ".mbox", ".rtf", ".md", ".markdown", ".odt",
    ".epub", ".docx", ".doc", ".pptx", ".ppt"
}

# ----------------------------------------------------------------------
# ZIP/RAR limits
# ----------------------------------------------------------------------
MAX_ZIP_SIZE = 100 * 1024 * 1024        # 100 MB
MAX_ZIP_FILES = 50
MAX_RAR_SIZE = 100 * 1024 * 1024
MAX_RAR_FILES = 50
# Per-member extraction limit to avoid oversized entries; set generously
MAX_ARCHIVE_MEMBER_SIZE = 100 * 1024 * 1024

# ----------------------------------------------------------------------
# Page/sheet limits
# ----------------------------------------------------------------------
MAX_PAGES_DEFAULT = 10
MAX_PAGES_ENV_VAR = "MAX_DOCUMENT_PAGES"
DISABLE_PAGE_LIMIT_ENV_VAR = "DISABLE_PAGE_LIMIT"

# ----------------------------------------------------------------------
# Excel direct-processing toggle (default: disabled)
# ----------------------------------------------------------------------
EXCEL_DIRECT_ENV_VAR = "ENABLE_DIRECT_EXCEL"

###############################################################################
# Logging configuration
###############################################################################
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

LOG_DIR = "processed_documents"
os.makedirs(LOG_DIR, exist_ok=True)
file_handler = logging.FileHandler(os.path.join(LOG_DIR, "processing.log"))
file_handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
logger.addHandler(file_handler)


###############################################################################
# Vision models
##############################################################################

load_dotenv()
DOCS_VISION_MODEL = os.getenv("DOCS_VISION_MODEL")
DOCS_VISION_MODEL_ROTATE = os.getenv("DOCS_VISION_MODEL_ROTATE")

###############################################################################
# Helper utilities
###############################################################################


def _safe_name(s: str) -> str:
    return "".join(c for c in s if c.isalnum() or c in ("_", "-"))


def _decode_mime_words(value: Optional[str]) -> str:
    if not value:
        return ""
    try:
        from email.header import decode_header
        parts = []
        for bytes_or_str, charset in decode_header(value):
            if isinstance(bytes_or_str, bytes):
                try:
                    parts.append(bytes_or_str.decode(charset or "utf-8", errors="ignore"))
                except Exception:
                    parts.append(bytes_or_str.decode("utf-8", errors="ignore"))
            else:
                parts.append(bytes_or_str)
        return "".join(parts).strip()
    except Exception:
        return str(value)


def _html_to_text(html: str) -> str:
    try:
        if BeautifulSoup is not None:
            soup = BeautifulSoup(html, "html.parser")
            return (soup.get_text("\n") or "").strip()
    except Exception:
        pass
    # Fallback: naive strip tags
    import re
    text = re.sub(r"<\s*br\s*/?\s*>", "\n", html, flags=re.I)
    text = re.sub(r"<[^>]+>", "", text)
    return text.strip()


def _generate_unique_image_name(source_type: str, original_name: str, extension: str) -> str:
    """
    Generate unique image filename to prevent conflicts across different processing types.
    {source_type}_{timestamp}_{counter}_{original_name}.{extension}
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    counter = getattr(_generate_unique_image_name, 'counter', 0) + 1
    _generate_unique_image_name.counter = counter
    clean_name = _safe_name(original_name).rstrip()
    return f"{source_type}_{timestamp}_{counter:03d}_{clean_name}.{extension}"


def _ensure_dirs() -> None:
    for d in ("images", "tables", LOG_DIR):
        os.makedirs(d, exist_ok=True)


def _save_binary(content: bytes, dest_path: str) -> None:
    os.makedirs(os.path.dirname(dest_path), exist_ok=True)
    with open(dest_path, "wb") as f:
        f.write(content)


def _save_image_data(data, dest_path: str) -> None:
    os.makedirs(os.path.dirname(dest_path), exist_ok=True)
    if hasattr(data, "save") and callable(getattr(data, "save")):
        data.save(dest_path)
    else:
        with open(dest_path, "wb") as f:
            f.write(data)


def _ensure_unique_path(path: str) -> str:
    if not os.path.exists(path):
        return path
    base, ext = os.path.splitext(path)
    i = 1
    while True:
        candidate = f"{base}_{i}{ext}"
        if not os.path.exists(candidate):
            return candidate
        i += 1


def _safe_join_path(root: str, arcname: str) -> str:
    # Normalize separators and strip dangerous parts
    rel = arcname.replace("\\", "/").lstrip("/")
    parts = []
    comps = rel.split("/")
    for i, part in enumerate(comps):
        if part in ("", "."):
            continue
        if part == "..":
            continue
        is_last = (i == len(comps) - 1)
        if is_last:
            # Keep dots in filename to preserve extension
            safe = "".join(c for c in part if c.isalnum() or c in ("_", "-", "."))
        else:
            safe = _safe_name(part)
        parts.append(safe)
    path = os.path.join(root, *parts) if parts else root
    # Ensure directory exists
    os.makedirs(os.path.dirname(path), exist_ok=True)
    return path


def _find_soffice() -> str:
    candidates = [
        os.getenv("SOFFICE_PATH"),
        shutil.which("soffice"),
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    ]
    for p in candidates:
        if p and os.path.isfile(p):
            return p
    raise FileNotFoundError(
        "LibreOffice 'soffice' executable not found. Install LibreOffice or set env var SOFFICE_PATH."
    )


@lru_cache(maxsize=1)
def _get_page_limit() -> Optional[int]:
    load_dotenv()
    disable_limit = os.getenv(DISABLE_PAGE_LIMIT_ENV_VAR, "").lower()
    if disable_limit in ("true", "1", "yes", "on"):
        logger.info("Page limits disabled via %s", DISABLE_PAGE_LIMIT_ENV_VAR)
        return None
    custom_limit = os.getenv(MAX_PAGES_ENV_VAR)
    if custom_limit:
        try:
            limit = int(custom_limit)
            if limit > 0:
                logger.info("Using custom page limit: %d", limit)
                return limit
            else:
                logger.warning("Invalid page limit %s, using default %d", custom_limit, MAX_PAGES_DEFAULT)
        except ValueError:
            logger.warning("Invalid page limit %s, using default %d", custom_limit, MAX_PAGES_DEFAULT)
    logger.info("Using default page limit: %d", MAX_PAGES_DEFAULT)
    return MAX_PAGES_DEFAULT


@lru_cache(maxsize=1)
def _is_direct_excel_enabled() -> bool:
    load_dotenv()
    flag = os.getenv(EXCEL_DIRECT_ENV_VAR, "").lower()
    return flag in ("true", "1", "yes", "on")


def _csv_previews(df: pd.DataFrame, label: str, *, preview_chars: int = CSV_PREVIEW_CHARS) -> List[str]:
    full_csv = df.to_csv(index=False)
    md = df.to_markdown(index=False)
    head = full_csv[:preview_chars]
    tail = "..." if len(full_csv) > preview_chars else ""
    return [
        f"{label} (CSV):\n{head}{tail}\n",
        f"{label} (Markdown):\n{md}\n",
    ]


def _get_api_key() -> Optional[str]:
    load_dotenv()
    return os.getenv("API_KEY") or os.getenv("OPENAI_API_KEY")

# --------------------------
# Safe Excel utilities
# --------------------------

def read_excel_file_safe(file_path: str, sheet_name: Optional[str] = None):
    file_ext = Path(file_path).suffix.lower()
    try:
        if file_ext == ".xlsx":
            try:
                return pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl")
            except Exception:
                return pd.read_excel(file_path, sheet_name=sheet_name)
        elif file_ext == ".xls":
            if not XLRD_AVAILABLE:
                raise ImportError("xlrd library is required for .xls files")
            workbook = xlrd.open_workbook(file_path)
            sheet = workbook.sheet_by_index(0) if sheet_name is None else workbook.sheet_by_name(sheet_name)
            data = []
            for row_idx in range(sheet.nrows):
                row_data = [sheet.cell(row_idx, col_idx).value for col_idx in range(sheet.ncols)]
                data.append(row_data)
            if data:
                df = pd.DataFrame(data[1:], columns=data[0]) if len(data) > 1 else pd.DataFrame(data)
            else:
                df = pd.DataFrame()
            return df
        else:
            return pd.read_excel(file_path, sheet_name=sheet_name)
    except Exception as e:
        logger.error("Error reading Excel file %s: %s", file_path, e)
        raise


def get_excel_sheet_names(file_path: str) -> List[str]:
    file_ext = Path(file_path).suffix.lower()
    try:
        if file_ext == ".xlsx":
            try:
                workbook = openpyxl.load_workbook(file_path, read_only=True)
                return workbook.sheetnames
            except Exception:
                with pd.ExcelFile(file_path) as xls:
                    return xls.sheet_names
        elif file_ext == ".xls":
            if not XLRD_AVAILABLE:
                logger.error("xlrd not available, cannot read .xls files")
                return []
            workbook = xlrd.open_workbook(file_path)
            return workbook.sheet_names()
        else:
            with pd.ExcelFile(file_path) as xls:
                return xls.sheet_names
    except Exception as e:
        logger.error("Error getting sheet names from %s: %s", file_path, e)
        return []


###############################################################################
# Document class
###############################################################################
class Document:
    _IMAGE_EXTS = SUPPORTED_IMAGE_FORMATS

    def __init__(self, file_path: str, *, media_dir: str = "media_for_processing") -> None:
        self.media_dir = media_dir
        os.makedirs(self.media_dir, exist_ok=True)
        if not file_path.startswith(self.media_dir):
            self.file_path = os.path.join(self.media_dir, os.path.basename(file_path))
        else:
            self.file_path = file_path
        self.file_name = os.path.basename(self.file_path)
        self.file_ext = os.path.splitext(self.file_name)[1].lower()
        # Normalize extension to strip trailing non-alnum chars like '}.eml' case
        if self.file_ext and not self.file_ext[-1].isalnum():
            # remove trailing non-alnum characters
            i = len(self.file_ext) - 1
            while i >= 0 and not self.file_ext[i].isalnum():
                i -= 1
            self.file_ext = self.file_ext[:i+1]
        self.file_size = os.path.getsize(self.file_path) if os.path.exists(self.file_path) else 0
        self.metadata: Dict[str, str] = {}
        self.text_content: str = ""
        self.tables: List[str] = []
        self.images: List[str] = []

    def process(self) -> None:
        if self.file_name.startswith("~$"):
            raise ValueError(f"Temporary file detected ('{self.file_name}'). Skipping processing.")
        _ensure_dirs()
        HANDLERS = {
            ".pdf": self._process_pdf,
            ".txt": self._process_txt,
            ".pages": self._process_pages,
            ".numbers": self._process_numbers,
            # spreadsheet handler guarded by env toggle
            # ".xlsx": self._process_spreadsheet,
            # ".xls": self._process_spreadsheet,
            ".csv": self._process_csv,
            ".json": self._process_generic_text,
            ".py": self._process_generic_text,
            ".html": self._process_generic_text,
            ".cms": self._process_generic_text,
            ".css": self._process_generic_text,
            ".eml": self._process_email,
            ".mbox": self._process_email,
            ".rtf": self._process_rtf,
            ".md": self._process_markdown,
            ".markdown": self._process_markdown,
            ".odt": self._process_odt,
            ".epub": self._process_epub,
            ".zip": self._process_generic_zip,
            ".rar": self._process_generic_rar,
        }
        ext = self.file_ext
        if ext in self._IMAGE_EXTS:
            handler = self._process_image
        elif ext in {".xlsx", ".xls"} and _is_direct_excel_enabled():
            handler = self._process_spreadsheet
        elif ext in {".docx", ".doc", ".pptx", ".ppt", ".xls", ".xlsx"}:
            pdf_path = self._convert_to_pdf(self.file_path)
            self.file_path, self.file_name, self.file_ext = pdf_path, os.path.basename(pdf_path), ".pdf"
            handler = self._process_pdf
        else:
            handler = HANDLERS.get(ext)
        if not handler:
            raise ValueError(f"Unsupported file format: {ext}")
        handler()

    # --------------------------
    # Format-specific methods
    # --------------------------

    def _process_pdf(self) -> None:
        logger.info("Processing PDF document %s", self.file_path)
        doc = fitz.open(self.file_path)
        self.metadata.update(doc.metadata or {})
        parts: List[str] = []
        page_limit = _get_page_limit()
        total_pages = len(doc)
        pages_to_process = min(total_pages, page_limit) if page_limit else total_pages
        if page_limit and total_pages > page_limit:
            # Align with spreadsheets behavior: mark that page limit was applied
            self.metadata['page_limit_reached'] = True
        for page_number in range(pages_to_process):
            page = doc[page_number]
            vision_calls = 0
            image_counter = 1
            page_dict = page.get_text("dict")
            blocks = page_dict.get("blocks", [])
            has_text = any(b.get("type") == 0 for b in blocks)
            has_images = any(b.get("type") == 1 for b in blocks)
            # Compute expensive HTML only if needed to disambiguate
            drawings_present = page.get_drawings()
            has_html_images = False
            if not drawings_present and not (has_images and not has_text):
                html_content = page.get_text("html")
                has_html_images = ("<img" in html_content and "base64" in html_content and
                                   len(html_content) > HTML_IMAGE_HEURISTIC_BYTES)
            if drawings_present or (has_images and not has_text) or has_html_images:
                logger.debug("Page %d has graphics (drawings=%s, only_images=%s, html_images=%s)",
                             page_number + 1, bool(drawings_present), bool(has_images and not has_text), bool(has_html_images))
                try:
                    pix = page.get_pixmap(dpi=200)
                    unique_name = _generate_unique_image_name("pdf", f"{os.path.splitext(self.file_name)[0]}_page{page_number + 1}", "png")
                    img_path = os.path.join("images", unique_name)
                    _save_image_data(pix, img_path)
                    self.images.append(img_path)
                    desc = self._generate_image_description(img_path)
                    parts.append(f"[========[Page {page_number + 1} with graphics]======== \n {desc}]\n")
                    logger.debug("Description for page %d with graphics added", page_number + 1)
                except Exception as e:
                    logger.error("Pixmap error p.%s: %s", page_number + 1, e)
                    logger.debug("Error processing page %d image with graphics", page_number + 1)
            else:
                logger.debug("Processing page %d: regular text + embedded raster images", page_number + 1)
                parts.append(f"========[Page {page_number + 1} regular text + embedded raster images]=======\n")
                if not blocks:
                    logger.debug("No text or images found on page %d", page_number + 1)
                for block in blocks:
                    if block.get("type") == 0:
                        for line in block.get("lines", []):
                            for span in line.get("spans", []):
                                parts.append(span.get("text", ""))
                    elif block.get("type") == 1:
                        parts.append(f"========[Image {image_counter}]========\n")
                        img_bytes = block.get("image")
                        w = block.get("width", 0)
                        h = block.get("height", 0)
                        if w * h < MIN_IMG_PIXELS:
                            continue
                        img_ext = block.get("ext", "png")
                        base_name = f"{os.path.splitext(self.file_name)[0]}_img{image_counter}"
                        unique_name = _generate_unique_image_name("pdf", base_name, img_ext)
                        img_path = os.path.join("images", unique_name)
                        try:
                            _save_image_data(img_bytes, img_path)
                            self.images.append(img_path)
                            if vision_calls < MAX_VISION_CALLS_PER_PAGE:
                                desc = self._generate_image_description(img_path)
                                vision_calls += 1
                            else:
                                desc = "(description skipped — Vision limit reached)"
                            parts.append(f"[Image {image_counter}: {desc}] (Image saved to: {img_path})")
                            image_counter += 1
                        except Exception as e:
                            logger.error("Save/describe image error: %s", e)
            parts.append("\n---\n")
        self.text_content = "".join(parts)
        doc.close()

    def _process_txt(self) -> None:
        with open(self.file_path, "r", encoding="utf-8", errors="ignore") as f:
            self.text_content = f.read()
        self._inject_basic_metadata()

    def _process_image(self) -> None:
        logger.info("Starting raster file processing: %s", self.file_path)
        try:
            img = Image.open(self.file_path)
        except Exception as e:
            raise ValueError(f"Cannot open image '{self.file_name}': {e}")
        try:
            exif = img.getexif()
            if exif and exif.get(274):
                img = ImageOps.exif_transpose(img)
        except Exception:
            pass
        if max(img.size) > MAX_IMAGE_DIM or self.file_size > MAX_IMAGE_SIZE:
            img.thumbnail((MAX_IMAGE_DIM, MAX_IMAGE_DIM))
        fmt = "JPEG" if img.format and img.format.lower() in {"tiff", "tif"} else (img.format or "PNG")
        base_name = os.path.splitext(self.file_name)[0]
        unique_name = _generate_unique_image_name("direct", base_name, fmt.lower())
        out_path = os.path.join("images", unique_name)
        _save_image_data(img, out_path)
        self.images.append(out_path)
        exif_data: Dict[str, str] = {}
        try:
            exif = img.getexif() or {}
            for tag, value in exif.items():
                exif_data[ExifTags.TAGS.get(tag, tag)] = value
            if "Orientation" in exif_data:
                exif_data["Orientation"] = 1
            self.metadata.update(exif_data)
        except Exception:
            logger.debug("No EXIF metadata or failed to parse for '%s'", self.file_name)
        self.text_content = f"{self._generate_image_description(out_path)} (Image saved to: {out_path})"

    def _process_spreadsheet(self) -> None:
        previews: List[str] = []
        sheet_names = get_excel_sheet_names(self.file_path)
        if not sheet_names:
            self.text_content = "No readable sheets found"
            self.metadata['error'] = 'no_sheets'
            return
        page_limit = _get_page_limit()
        total_sheets = len(sheet_names)
        sheets_to_process = min(total_sheets, page_limit) if page_limit else total_sheets
        if page_limit and total_sheets > page_limit:
            logger.info("Excel has %d sheets, limiting processing to %d sheets", total_sheets, page_limit)
            self.metadata['page_limit_reached'] = True
        base = os.path.splitext(self.file_name)[0]
        for sheet_name in sheet_names[:sheets_to_process]:
            try:
                df = read_excel_file_safe(self.file_path, sheet_name=sheet_name)
                if df.empty:
                    continue
                csv_path = os.path.join("tables", f"{base}_{sheet_name}.csv")
                os.makedirs("tables", exist_ok=True)
                df.to_csv(csv_path, index=False)
                self.tables.append(csv_path)
                previews.extend(_csv_previews(df, f"Sheet: {sheet_name}"))
            except Exception as e:
                logger.error("Error processing sheet %s: %s", sheet_name, e)
                previews.append(f"Sheet: {sheet_name} (ERROR)\nError: {str(e)}\n")
        self.text_content = "\n".join(previews)

    def _process_csv(self) -> None:
        df = pd.read_csv(self.file_path)
        csv_copy = os.path.join("tables", os.path.basename(self.file_path))
        df.to_csv(csv_copy, index=False)
        self.tables.append(csv_copy)
        full_csv = df.to_csv(index=False)
        md = df.to_markdown(index=False)
        self.text_content = (f"CSV (full):\n{full_csv}\n" f"CSV (Markdown):\n{md}")

    def _process_generic_text(self) -> None:
        with open(self.file_path, "r", encoding="utf-8", errors="ignore") as f:
            self.text_content = f.read()
        self._inject_basic_metadata()

    def _process_rtf(self) -> None:
        with open(self.file_path, "r", encoding="utf-8", errors="ignore") as f:
            data = f.read()
        self.text_content = rtf_to_text(data)
        self._inject_basic_metadata()

    def _process_markdown(self) -> None:
        with open(self.file_path, "r", encoding="utf-8", errors="ignore") as f:
            md = f.read()
        self.text_content = strip_markdown.strip_markdown(md)
        self._inject_basic_metadata()

    def _process_odt(self) -> None:
        with zipfile.ZipFile(self.file_path, "r") as z:
            if "content.xml" not in z.namelist():
                raise ValueError("ODT missing content.xml")
            raw_xml = z.read("content.xml").decode("utf-8", errors="ignore")
        import re
        text = re.sub(r"<[^>]+>", "", raw_xml.replace("</text:p>", "\n"))
        self.text_content = text.strip()
        self._inject_basic_metadata()

    def _process_epub(self) -> None:
        doc = fitz.open(self.file_path)
        pages = [p.get_text("text") for p in doc]
        self.text_content = "\n".join(pages)
        doc.close()
        self._inject_basic_metadata()

    def _process_pages(self) -> None:
        self._process_zip_bundle(expect_xml=True)

    def _process_numbers(self) -> None:
        """
        Process Apple Numbers (.numbers) files using numbers-parser when available.
        - Extract tables from all sheets, exporting to CSV in `tables/`.
        - Populate text_content with CSV/Markdown previews.
        - Robust exception handling with fallback to legacy zip-bundle logic.
        """
        # Fast-fail to legacy behavior if library not present
        if not NUMBERS_PARSER_AVAILABLE:
            logger.info("numbers-parser not available; falling back to zip-bundle logic for %s", self.file_name)
            try:
                self._process_zip_bundle(expect_xml=True)
            except Exception as e:
                logger.error("Zip-bundle fallback failed for %s: %s", self.file_name, e)
                self.text_content = "(failed to process .numbers: parser unavailable and zip fallback failed)"
                self.metadata['error'] = 'numbers_parser_unavailable'
            return

        previews: List[str] = []
        base_name = os.path.splitext(self.file_name)[0]
        page_limit = _get_page_limit()
        try:
            doc = NumbersDocument(self.file_path)
            sheets = list(getattr(doc, 'sheets', []))
            total_sheets = len(sheets)
            sheets_to_process = min(total_sheets, page_limit) if page_limit else total_sheets
            if page_limit and total_sheets > page_limit:
                logger.info("Numbers has %d sheets, limiting to %d", total_sheets, sheets_to_process)
                self.metadata['page_limit_reached'] = True

            for sheet in sheets[:sheets_to_process]:
                sheet_name = getattr(sheet, 'name', 'Sheet')
                tables = list(getattr(sheet, 'tables', []))
                if not tables:
                    continue
                for table in tables:
                    table_name = getattr(table, 'name', 'Table')
                    # Extract rows; numbers-parser returns Cell objects; coerce to values
                    raw_rows = list(table.rows()) if hasattr(table, 'rows') else []
                    value_rows = []
                    for row in raw_rows:
                        # Row may be list of Cell or primitives
                        values = []
                        for cell in row:
                            try:
                                values.append(getattr(cell, 'value', cell))
                            except Exception:
                                values.append(str(cell))
                        value_rows.append(values)
                    if not value_rows:
                        continue
                    headers = value_rows[0]
                    data_rows = value_rows[1:] if len(value_rows) > 1 else []
                    # Ensure headers are strings and unique-ish
                    str_headers: List[str] = []
                    seen = {}
                    for idx, h in enumerate(headers):
                        key = str(h) if h is not None and str(h).strip() != '' else f"col_{idx+1}"
                        if key in seen:
                            seen[key] += 1
                            key = f"{key}_{seen[key]}"
                        else:
                            seen[key] = 1
                        str_headers.append(key)

                    try:
                        os.makedirs("tables", exist_ok=True)
                        safe_sheet = _safe_name(str(sheet_name))
                        safe_table = _safe_name(str(table_name))
                        csv_path = os.path.join("tables", f"{base_name}_{safe_sheet}_{safe_table}.csv")
                        df = pd.DataFrame(data_rows, columns=str_headers)
                        df.to_csv(csv_path, index=False)
                        self.tables.append(csv_path)
                        previews.extend(_csv_previews(df, f"Sheet: {sheet_name} / Table: {table_name}"))
                    except Exception as e:
                        logger.error("Failed exporting sheet '%s' table '%s' to CSV: %s", sheet_name, table_name, e)
                        previews.append(
                            f"Sheet: {sheet_name} / Table: {table_name} (ERROR)\nError: {str(e)}\n"
                        )

            if previews:
                self.text_content = "\n".join(previews)
            else:
                self.text_content = "No readable tables found"
            self._inject_basic_metadata()
        except FileNotFoundError:
            logger.error(".numbers file not found: %s", self.file_path)
            self.metadata['error'] = 'numbers_file_not_found'
            self.text_content = "(.numbers file not found)"
        except PermissionError:
            logger.error("Permission denied reading .numbers: %s", self.file_path)
            self.metadata['error'] = 'numbers_permission_denied'
            self.text_content = "(permission denied reading .numbers)"
        except Exception as e:
            logger.error("numbers-parser processing failed for %s: %s", self.file_name, e)
            self.metadata['error'] = 'numbers_parser_failed'
            # Fallback to legacy zip-based minimal extraction (images/xml)
            try:
                self._process_zip_bundle(expect_xml=True)
                # annotate that fallback occurred
                self.metadata['fallback'] = 'zip_bundle'
            except Exception as e2:
                logger.error("Zip-bundle fallback also failed for %s: %s", self.file_name, e2)
                self.text_content = "(failed to process .numbers: parser and zip fallback failed)"

    # --------------------------
    # Helpers
    # --------------------------

    def _inject_basic_metadata(self) -> None:
        stat = os.stat(self.file_path)
        self.metadata.update({
            "modified_time": datetime.fromtimestamp(stat.st_mtime).isoformat(),
            "created_time": datetime.fromtimestamp(stat.st_ctime).isoformat(),
            "size_bytes": stat.st_size,
        })

    def _process_zip_bundle(self, *, expect_xml: bool = False) -> None:
        with zipfile.ZipFile(self.file_path, "r") as z:
            xml_files = [f for f in z.namelist() if f.endswith(".xml")]
            jpg_files = [f for f in z.namelist() if os.path.splitext(f)[1].lower() in SUPPORTED_IMAGE_FORMATS]
            if jpg_files:
                data = z.read(jpg_files[0])
                original_name = os.path.splitext(os.path.basename(jpg_files[0]))[0]
                extension = os.path.splitext(jpg_files[0])[1][1:]
                unique_name = _generate_unique_image_name("archive", original_name, extension)
                img_path = os.path.join("images", unique_name)
                _save_binary(data, img_path)
                self.images.append(img_path)
                self.text_content += f"{self._generate_image_description(img_path)} (Image saved to: {img_path})\n"
            if expect_xml and xml_files:
                with z.open(xml_files[0]) as f:
                    self.text_content += f.read().decode("utf-8", errors="ignore")
            elif expect_xml and not xml_files:
                self.text_content += "(XML not found in bundle)"
        self._inject_basic_metadata()

    def _convert_to_pdf(self, path: str) -> str:
        soffice = _find_soffice()
        output_dir = os.path.dirname(path)
        subprocess.run([soffice, "--headless", "--convert-to", "pdf", "--outdir", output_dir, path], check=True)
        pdf_path = os.path.splitext(path)[0] + ".pdf"
        logger.debug("Converted '%s' to PDF via %s", path, soffice)
        return pdf_path

    # --------------------------
    # OpenAI Vision
    # --------------------------

    def _fix_image_orientation(self, b64_string: str, *, model: str = DOCS_VISION_MODEL_ROTATE, timeout: float = 180) -> str:
        """
        Detect image orientation from base64, rotate in-memory, return new base64.
        Returns original base64 on error or angle=0.
        """
        api_key = _get_api_key()
        if not api_key:
            logger.warning("API_KEY not set; skipping image orientation fix.")
            return b64_string
        client = OpenAI(api_key=api_key)
        try:
            completion = client.chat.completions.create(
                model=model,
                response_format={"type": "json_object"},
                messages=[
                    {"role": "system", "content": "You are a helpful assistant that detects image orientation and responds in JSON."},
                    {"role": "user", "content": [
                        {"type": "text", "text": "Detect the clockwise rotation (0, 90, 180, 270) needed to make the document upright. Return {\"angle\": <int>}"},
                        {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64_string}"}},
                    ]},
                ],
                max_tokens=100,
                temperature=0,
                timeout=timeout,
            )
            response_content = completion.choices[0].message.content
            angle_json = json.loads(response_content)
            angle = angle_json.get("angle")
            if angle in {90, 180, 270}:
                image_data = base64.b64decode(b64_string)
                img = Image.open(io.BytesIO(image_data))
                rotated_img = img.rotate(-angle, expand=True)
                buffered = io.BytesIO()
                img_format = img.format or "JPEG"
                rotated_img.save(buffered, format=img_format)
                return base64.b64encode(buffered.getvalue()).decode()
            return b64_string
        except Exception as e:
            logger.error("OpenAI orientation detection failed: %s", e)
            return b64_string

    def _generate_image_description(self, image_path: str, *, model: str = DOCS_VISION_MODEL, max_tokens: int = 5000, timeout: float = 180) -> str:
        api_key = _get_api_key()
        if not api_key:
            logger.warning("API_KEY not set; skipping Vision description.")
            return "(description unavailable)"
        client = OpenAI(api_key=api_key)
        with open(image_path, "rb") as img_file:
            b64 = base64.b64encode(img_file.read()).decode()
        # Correct orientation before sending to description
        b64 = self._fix_image_orientation(b64_string=b64)
        try:
            completion = client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "user", "content": [
                        {"type": "text", "text": AI_IMAGE_DESCRIPTION_PROMPT},
                        {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64}"}},
                    ]}
                ],
                max_tokens=max_tokens,
                temperature=0,
                timeout=timeout,
            )
            return completion.choices[0].message.content.strip()
        except Exception as e:
            logger.error("OpenAI image description failed: %s", e)
            return "(description failed)"

    def _process_email(self) -> None:
        # Parse .eml, extract headers/body, save attachments and process them
        headers_out: List[str] = []
        body_text: str = ""
        attachments_info: List[str] = []
        attachments_processed = 0
        try:
            with open(self.file_path, "rb") as f:
                msg = BytesParser(policy=policy.default).parse(f)
        except Exception as e:
            self.text_content = f"(failed to parse email: {e})"
            self.metadata['error'] = 'email_parse_failed'
            return

        # Headers
        hdr_map = {
            "From": _decode_mime_words(msg.get("From")),
            "To": _decode_mime_words(msg.get("To")),
            "Cc": _decode_mime_words(msg.get("Cc")),
            "Bcc": _decode_mime_words(msg.get("Bcc")),
            "Date": _decode_mime_words(msg.get("Date")),
            "Subject": _decode_mime_words(msg.get("Subject")),
            "Message-ID": _decode_mime_words(msg.get("Message-ID")),
        }
        for k, v in hdr_map.items():
            if v:
                headers_out.append(f"{k}: {v}")

        # Walk parts to find body and attachments
        plain_found = False
        html_candidate = None
        attachments_dir = os.path.join(self.media_dir, "attachments", _safe_name(os.path.splitext(self.file_name)[0]))
        os.makedirs(attachments_dir, exist_ok=True)

        if msg.is_multipart():
            for part in msg.walk():
                ctype = part.get_content_type()
                disp = (part.get_content_disposition() or "").lower()
                filename = part.get_filename()
                if filename:
                    filename = _decode_mime_words(filename)
                if disp in {"attachment", "inline"} and filename:
                    try:
                        data = part.get_payload(decode=True) or b""
                        safe_base = _safe_name(os.path.splitext(filename)[0]) or "attachment"
                        ext = os.path.splitext(filename)[1] or ""
                        dest = os.path.join(attachments_dir, f"{safe_base}{ext}")
                        dest = _ensure_unique_path(dest)
                        with open(dest, "wb") as out:
                            out.write(data)
                        attachments_processed += 1
                        attachments_info.append(f"Saved attachment: {os.path.basename(dest)}")
                        # Process attachment with Document
                        try:
                            child_doc = Document(dest, media_dir=self.media_dir)
                            child_doc.process()
                            self.images.extend(child_doc.images)
                            if hasattr(child_doc, 'tables'):
                                self.tables.extend(child_doc.tables)
                            attachments_info.append(f"Attachment processed: {os.path.basename(dest)}")
                            # Append a short preview of child content
                            preview = child_doc.text_content[:500]
                            body_text += f"\n---\n[Attachment content preview: {os.path.basename(dest)}]\n{preview}\n"
                        except Exception as ce:
                            attachments_info.append(f"Attachment processing failed: {os.path.basename(dest)} ({ce})")
                    except Exception as se:
                        attachments_info.append(f"Attachment save failed: {filename} ({se})")
                    continue
                # Body parts
                if ctype == "text/plain" and not plain_found:
                    try:
                        body_text = part.get_content() or ""
                        plain_found = True
                    except Exception:
                        pass
                elif ctype == "text/html" and not plain_found and html_candidate is None:
                    try:
                        html_candidate = part.get_content() or ""
                    except Exception:
                        pass
        else:
            # Singlepart
            ctype = msg.get_content_type()
            try:
                if ctype == "text/plain":
                    body_text = msg.get_content() or ""
                    plain_found = True
                elif ctype == "text/html":
                    html_candidate = msg.get_content() or ""
            except Exception:
                pass

        if not plain_found and html_candidate:
            body_text = _html_to_text(html_candidate)

        # Assemble text_content
        lines: List[str] = []
        lines.append("==== Email Headers ====")
        lines.extend(headers_out)
        lines.append("\n==== Email Body ====")
        lines.append(body_text or "(no body)")
        if attachments_info:
            lines.append("\n==== Attachments ====")
            lines.extend(attachments_info)
        self.text_content = "\n".join(lines)
        if attachments_processed:
            self.metadata['attachments_count'] = attachments_processed
        self._inject_basic_metadata()

    # --------------------------
    # Generic ZIP/RAR
    # --------------------------

    def _process_generic_zip(self) -> None:
        logger.info("Processing ZIP archive '%s' (%d bytes)", self.file_name, self.file_size)
        if self.file_size > MAX_ZIP_SIZE:
            msg = (f"(archive skipped: size {self.file_size} B > {MAX_ZIP_SIZE} B limit)")
            logger.warning("%s %s", self.file_name, msg)
            self.text_content = msg
            return
        extracted_root = os.path.join(self.media_dir, "unzipped", os.path.splitext(self.file_name)[0])
        os.makedirs(extracted_root, exist_ok=True)
        useful_files = 0
        try:
            with zipfile.ZipFile(self.file_path, "r") as z:
                names = z.namelist()
                if len(names) > MAX_ZIP_FILES:
                    logger.info("ZIP %s has %d entries; only first %d will be processed", self.file_name, len(names), MAX_ZIP_FILES)
                for idx, name in enumerate(names):
                    if idx >= MAX_ZIP_FILES:
                        break
                    if name.endswith("/"):
                        continue
                    _, ext = os.path.splitext(name)
                    ext = ext.lower()
                    if ext in ARCHIVE_ALLOWED_EXTS:
                        try:
                            info = z.getinfo(name)
                            if info.file_size > MAX_ARCHIVE_MEMBER_SIZE:
                                msg = (f"(member skipped due to size limit: {name} size {info.file_size} B > "
                                       f"{MAX_ARCHIVE_MEMBER_SIZE} B)")
                                logger.warning("%s %s", self.file_name, msg)
                                self.text_content += f"\n{msg}\n"
                                continue
                        except KeyError:
                            # Fallback if info missing; proceed
                            pass
                        dest_path = _safe_join_path(extracted_root, name)
                        if os.path.exists(dest_path):
                            dest_path = _ensure_unique_path(dest_path)
                        with z.open(name) as src, open(dest_path, "wb") as dst:
                            shutil.copyfileobj(src, dst)
                        logger.info("Extracted %s from %s", name, self.file_name)
                        try:
                            child_doc = Document(dest_path, media_dir=self.media_dir)
                            child_doc.process()
                            useful_files += 1
                            self.text_content += (f"\n===== [Extracted: {name}] =====\n" f"{child_doc.text_content}\n")
                            self.images.extend(child_doc.images)
                            if hasattr(child_doc, "tables"):
                                self.tables.extend(child_doc.tables)
                        except Exception as e:
                            logger.error("Failed processing '%s' inside '%s': %s", name, self.file_name, e)
                    else:
                        logger.debug("Skipping unsupported entry '%s' in %s", name, self.file_name)
        except Exception as e:
            logger.error("Cannot open ZIP '%s': %s", self.file_name, e)
            self.text_content = "(zip archive corrupted or unreadable)"
            return
        if useful_files == 0:
            self.text_content = "(zip contains no supported documents or images)"
        self._inject_basic_metadata()

    def _process_generic_rar(self) -> None:
        if rarfile is None:
            msg = "(rarfile module not installed; .rar skipped)"
            logger.warning("%s %s", self.file_name, msg)
            self.text_content = msg
            return
        logger.info("Processing RAR archive '%s' (%d bytes)", self.file_name, self.file_size)
        if self.file_size > MAX_RAR_SIZE:
            msg = (f"(archive skipped: size {self.file_size} B > {MAX_RAR_SIZE} B limit)")
            logger.warning("%s %s", self.file_name, msg)
            self.text_content = msg
            return
        extracted_root = os.path.join(self.media_dir, "unrarred", os.path.splitext(self.file_name)[0])
        os.makedirs(extracted_root, exist_ok=True)
        useful_files = 0
        try:
            with rarfile.RarFile(self.file_path) as rf:
                # Prefer detailed infos if available
                if hasattr(rf, 'infolist'):
                    entries = [(i.filename, getattr(i, 'file_size', 0)) for i in rf.infolist()]
                else:
                    names = rf.namelist()
                    entries = [(n, 0) for n in names]
                if len(entries) > MAX_RAR_FILES:
                    logger.info("RAR %s has %d entries; only first %d will be processed", self.file_name, len(entries), MAX_RAR_FILES)
                for idx, (name, fsize) in enumerate(entries):
                    if idx >= MAX_RAR_FILES:
                        break
                    if name.endswith("/"):
                        continue
                    _, ext = os.path.splitext(name)
                    ext = ext.lower()
                    if ext in ARCHIVE_ALLOWED_EXTS:
                        try:
                            size_ok = True
                            if fsize and fsize > MAX_ARCHIVE_MEMBER_SIZE:
                                size_ok = False
                            if not size_ok:
                                msg = (f"(member skipped due to size limit: {name} size {fsize} B > "
                                       f"{MAX_ARCHIVE_MEMBER_SIZE} B)")
                                logger.warning("%s %s", self.file_name, msg)
                                self.text_content += f"\n{msg}\n"
                                continue
                        except Exception:
                            pass
                        dest_path = _safe_join_path(extracted_root, name)
                        if os.path.exists(dest_path):
                            dest_path = _ensure_unique_path(dest_path)
                        with rf.open(name) as src, open(dest_path, "wb") as dst:
                            shutil.copyfileobj(src, dst)
                        logger.info("Extracted %s from %s", name, self.file_name)
                        try:
                            child_doc = Document(dest_path, media_dir=self.media_dir)
                            child_doc.process()
                            useful_files += 1
                            self.text_content += (f"\n===== [Extracted: {name}] =====\n" f"{child_doc.text_content}\n")
                            self.images.extend(child_doc.images)
                            if hasattr(child_doc, "tables"):
                                self.tables.extend(child_doc.tables)
                        except Exception as e:
                            logger.error("Failed processing '%s' inside '%s': %s", name, self.file_name, e)
                    else:
                        logger.debug("Skipping unsupported entry '%s' in %s", name, self.file_name)
        except Exception as e:
            # Try to classify common rarfile exceptions for better UX
            msg = None
            try:
                PR = getattr(rarfile, 'PasswordRequired', None)
                BR = getattr(rarfile, 'BadRarFile', None)
                if PR and isinstance(e, PR):
                    msg = "(rar is password-protected; skipped)"
                elif BR and isinstance(e, BR):
                    msg = "(rar archive corrupted or unreadable)"
            except Exception:
                pass
            if msg is None:
                msg = f"(rar open failed: {e})"
            logger.error("%s: %s", self.file_name, msg)
            self.text_content = msg
            return
        if useful_files == 0:
            self.text_content = "(rar contains no supported documents or images)"
        self._inject_basic_metadata()


###############################################################################
# Batch utility
###############################################################################

def batch_process_folder(
    input_folder: str,
    output_file: str = None,
    *,
    preview_chars: int = None,
    return_docs: bool = False,
) -> list:
    _ensure_dirs()
    docs = []  # type: List[dict]
    results: List[str] = []
    if not return_docs:
        results.append("=== Batch processing report ===\n")
    for root, _, files in os.walk(input_folder):
        for name in sorted(files):
            if name.startswith("."):
                continue
            abs_path = os.path.join(root, name)
            rel_path = os.path.relpath(abs_path, input_folder)
            dest_path = os.path.join("media_for_processing", rel_path)
            os.makedirs(os.path.dirname(dest_path), exist_ok=True)
            if not os.path.exists(dest_path):
                shutil.copy2(abs_path, dest_path)
            if not return_docs:
                results.append("======================================")
                results.append(f"File: {rel_path}")
            try:
                doc = Document(dest_path)
                doc.process()
                doc_info = {
                    'rel_path': rel_path,
                    'file_ext': doc.file_ext,
                    'file_size': doc.file_size,
                    'metadata': doc.metadata,
                    'text_content': doc.text_content,
                    'absolute_path': dest_path,
                    'tables': getattr(doc, 'tables', []),
                }
                docs.append(doc_info)
                if not return_docs:
                    results.append(f"Type: {doc.file_ext}")
                    results.append(f"Size: {doc.file_size} bytes")
                    results.append("----------- Metadata -----------")
                    results.extend(f"{k}: {v}" for k, v in doc.metadata.items())
                    results.append("----------- Content -----------")
                    if preview_chars is None:
                        results.append(doc.text_content)
                    else:
                        txt = doc.text_content
                        results.append(txt[:preview_chars] + ("..." if len(txt) > preview_chars else ""))
            except Exception as e:
                results.append(f"ERROR: {e}")
            if not return_docs:
                results.append("======================================\n")
    if return_docs:
        return docs
    if output_file:
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        with open(output_file, "w", encoding="utf-8") as f:
            f.write("\n".join(results))
        print(f"Batch processing completed. Report at {output_file}")
    return docs


if __name__ == "__main__":
    batch_process_folder(
        "downloaded_files",
        os.path.join("processed_documents_21_04", "final_output_21_04_94.txt"),
    )
